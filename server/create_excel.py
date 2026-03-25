import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ─── SHEET 1: Players ───────────────────────────────────────────────
ws_players = wb.active
ws_players.title = "Players"

header_fill = PatternFill("solid", start_color="1a1a2e")
header_font = Font(bold=True, color="e94560", size=11, name="Arial")
cell_font = Font(name="Arial", size=10)

players_headers = ["PlayerName", "SecretCode", "IsActive"]
for col, h in enumerate(players_headers, 1):
    c = ws_players.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

players = [
    ("Athena",   "ATHENA2026",  True),
    ("Stuti",    "STUTI2026",   True),
    ("Vindhya",  "VINDHYA2026", True),
    ("Sanika",   "SANIKA2026",  True),
    ("Kulvir",   "KULVIR2026",  True),
    ("Himanshu", "HIMANSHU26",  True),
    ("Soham",    "SOHAM2026",   True),
    ("Sahil",    "SAHIL2026",   True),
    ("Vidit",    "VIDIT2026",   True),
    ("MODERATOR","MOD_IPL2026", True),
]
for row, (name, code, active) in enumerate(players, 2):
    ws_players.cell(row=row, column=1, value=name).font = cell_font
    ws_players.cell(row=row, column=2, value=code).font = cell_font
    ws_players.cell(row=row, column=3, value=active).font = cell_font

ws_players.column_dimensions["A"].width = 15
ws_players.column_dimensions["B"].width = 18
ws_players.column_dimensions["C"].width = 12

# ─── SHEET 2: Fixtures ──────────────────────────────────────────────
ws_fix = wb.create_sheet("Fixtures")

fix_headers = [
    "MatchNo", "Date", "Day", "Team1", "Team2", "Venue",
    "Time", "Status", "ActualWinner", "ActualMOM"
]
for col, h in enumerate(fix_headers, 1):
    c = ws_fix.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

fixtures = [
    (1,  "28-Mar-2026", "Sat", "RCB",  "SRH",  "Bengaluru",      "7:30 PM"),
    (2,  "29-Mar-2026", "Sun", "MI",   "KKR",  "Mumbai",         "7:30 PM"),
    (3,  "30-Mar-2026", "Mon", "RR",   "CSK",  "Guwahati",       "7:30 PM"),
    (4,  "31-Mar-2026", "Tue", "PBKS", "GT",   "New Chandigarh", "7:30 PM"),
    (5,  "01-Apr-2026", "Wed", "LSG",  "DC",   "Lucknow",        "7:30 PM"),
    (6,  "02-Apr-2026", "Thu", "KKR",  "SRH",  "Kolkata",        "7:30 PM"),
    (7,  "03-Apr-2026", "Fri", "CSK",  "PBKS", "Chennai",        "7:30 PM"),
    (8,  "04-Apr-2026", "Sat", "DC",   "MI",   "Delhi",          "3:30 PM"),
    (9,  "04-Apr-2026", "Sat", "GT",   "RR",   "Ahmedabad",      "7:30 PM"),
    (10, "05-Apr-2026", "Sun", "SRH",  "LSG",  "Hyderabad",      "3:30 PM"),
    (11, "05-Apr-2026", "Sun", "RCB",  "CSK",  "Bengaluru",      "7:30 PM"),
    (12, "06-Apr-2026", "Mon", "KKR",  "PBKS", "Kolkata",        "7:30 PM"),
    (13, "07-Apr-2026", "Tue", "RR",   "MI",   "Guwahati",       "7:30 PM"),
    (14, "08-Apr-2026", "Wed", "DC",   "GT",   "Delhi",          "7:30 PM"),
    (15, "09-Apr-2026", "Thu", "KKR",  "LSG",  "Kolkata",        "7:30 PM"),
    (16, "10-Apr-2026", "Fri", "RR",   "RCB",  "Guwahati",       "7:30 PM"),
    (17, "11-Apr-2026", "Sat", "PBKS", "SRH",  "New Chandigarh", "3:30 PM"),
    (18, "11-Apr-2026", "Sat", "CSK",  "DC",   "Chennai",        "7:30 PM"),
    (19, "12-Apr-2026", "Sun", "LSG",  "GT",   "Lucknow",        "3:30 PM"),
    (20, "12-Apr-2026", "Sun", "MI",   "RCB",  "Mumbai",         "7:30 PM"),
]

alt_fill = PatternFill("solid", start_color="0f3460")
for row, f in enumerate(fixtures, 2):
    fill = alt_fill if row % 2 == 0 else PatternFill("solid", start_color="16213e")
    for col, val in enumerate(f, 1):
        c = ws_fix.cell(row=row, column=col, value=val)
        c.font = cell_font
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
    ws_fix.cell(row=row, column=8, value="UPCOMING")
    ws_fix.cell(row=row, column=9, value="")
    ws_fix.cell(row=row, column=10, value="")

col_widths = [10, 15, 8, 8, 8, 18, 10, 12, 15, 20]
for i, w in enumerate(col_widths, 1):
    ws_fix.column_dimensions[get_column_letter(i)].width = w

# ─── SHEET 3: Predictions ───────────────────────────────────────────
ws_pred = wb.create_sheet("Predictions")

pred_headers = ["MatchNo", "PlayerName", "PredictedWinner", "PredictedMOM", "SubmittedAt", "WinnerPoints", "MOMPoints", "TotalPoints"]
for col, h in enumerate(pred_headers, 1):
    c = ws_pred.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

# Pre-populate all match/player combos with empty predictions
row = 2
for match_no in range(1, 21):
    for pname, _, _ in players[:9]:  # exclude moderator
        ws_pred.cell(row=row, column=1, value=match_no)
        ws_pred.cell(row=row, column=2, value=pname)
        ws_pred.cell(row=row, column=3, value="")
        ws_pred.cell(row=row, column=4, value="")
        ws_pred.cell(row=row, column=5, value="")
        ws_pred.cell(row=row, column=6, value=0)
        ws_pred.cell(row=row, column=7, value=0)
        ws_pred.cell(row=row, column=8, value=0)
        for col in range(1, 9):
            ws_pred.cell(row=row, column=col).font = cell_font
        row += 1

pred_widths = [10, 15, 18, 25, 22, 15, 12, 14]
for i, w in enumerate(pred_widths, 1):
    ws_pred.column_dimensions[get_column_letter(i)].width = w

wb.save("/home/claude/ipl-fantasy/server/ipl_fantasy.xlsx")
print("Excel created successfully!")
